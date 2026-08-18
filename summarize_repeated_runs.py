#!/usr/bin/env python3
"""Aggregate per-run training summary workbooks as mean +/- sample standard deviation."""

from __future__ import annotations

import argparse
import datetime as dt
import os
import statistics
from collections import defaultdict
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SUMMARY_START = "=== REPEATED RUN SUMMARY ==="
SUMMARY_END = "=== END REPEATED RUN SUMMARY ==="

METRICS = [
    ("loss", "last_val_loss", "workload_loss"),
    ("mse", "last_val_mse", "workload_mse"),
    ("mape", "last_val_mape", "workload_mape"),
    ("q50", "last_val_q50", "workload_q50"),
    ("q95", "last_val_q95", "workload_q95"),
    ("q99", "last_val_q99", "workload_q99"),
    ("worst_q", None, "workload_worst_q"),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("summary_xlsx", nargs="+", help="Per-run summary XLSX files")
    parser.add_argument("--requested-runs", type=int, required=True)
    parser.add_argument("--output-xlsx", required=True, help="Aggregate XLSX output path")
    return parser.parse_args()


def worksheet_rows(worksheet) -> List[Dict[str, object]]:
    values = worksheet.iter_rows(values_only=True)
    headers = next(values, None)
    if not headers:
        return []
    return [dict(zip(headers, row)) for row in values]


def read_workbooks(paths: Iterable[str]):
    rows: List[Dict[str, object]] = []
    run_details: List[Dict[str, object]] = []
    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "Results" not in workbook.sheetnames:
            raise ValueError(f"Results sheet missing from {path}")
        workbook_rows = worksheet_rows(workbook["Results"])
        for row in workbook_rows:
            row["_summary_xlsx"] = path
            rows.append(row)

        details = {"summary_xlsx": path}
        if "Configuration" in workbook.sheetnames:
            for item in worksheet_rows(workbook["Configuration"]):
                source = str(item.get("source") or "configuration")
                variable = str(item.get("variable") or "")
                if variable:
                    details[f"{source}.{variable}"] = item.get("value")
        run_details.append(details)
        workbook.close()
    return rows, run_details


def as_float(value: object) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def mean_std(values: Iterable[Optional[float]]) -> Tuple[Optional[float], Optional[float], int]:
    clean = [value for value in values if value is not None]
    if not clean:
        return None, None, 0
    std = statistics.stdev(clean) if len(clean) > 1 else 0.0
    return statistics.mean(clean), std, len(clean)


def fmt_pair(pair: Tuple[Optional[float], Optional[float], int]) -> str:
    mean, std, _ = pair
    if mean is None or std is None:
        return "-"
    return f"{mean:.4f} ± {std:.4f}"


def unique_join(rows: List[Dict[str, object]], key: str) -> str:
    values = sorted({str(row.get(key, "")) for row in rows if row.get(key, "")})
    return ", ".join(values) if values else "-"


def aggregate(rows: List[Dict[str, object]], requested_runs: int, run_details=None):
    rows_by_log: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        rows_by_log[row.get("log_file") or row["_summary_xlsx"]].append(row)

    successful_logs = {
        log_file for log_file, log_rows in rows_by_log.items()
        if log_rows and log_rows[0].get("run_status") == "SUCCESS"
    }
    successful_rows = [row for row in rows if (row.get("log_file") or row["_summary_xlsx"]) in successful_logs]
    run_rows = [rows_by_log[log_file][0] for log_file in sorted(successful_logs)]

    validation = {}
    for label, val_key, _ in METRICS:
        if val_key is not None:
            validation[label] = mean_std(as_float(row.get(val_key)) for row in run_rows)

    workload_rows: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in successful_rows:
        workload = str(row.get("workload") or "").strip()
        if workload:
            workload_rows[workload].append(row)

    workloads = {}
    for workload, grouped_rows in sorted(workload_rows.items()):
        workloads[workload] = {
            label: mean_std(as_float(row.get(workload_key)) for row in grouped_rows)
            for label, _, workload_key in METRICS
        }

    return {
        "requested_runs": requested_runs,
        "successful_runs": len(successful_logs),
        "failed_runs": max(requested_runs - len(successful_logs), 0),
        "run_rows": run_rows,
        "validation": validation,
        "workloads": workloads,
        "run_details": run_details or [],
    }


def print_summary(summary) -> None:
    run_rows = summary["run_rows"]
    print()
    print(SUMMARY_START)
    print(f"Generated at: {dt.datetime.now().astimezone().strftime('%Y-%m-%d %H:%M:%S %Z')}")
    print(f"Runs: {summary['successful_runs']}/{summary['requested_runs']} successful")
    print(f"CUDA device(s): {unique_join(run_rows, 'cuda_device')}")
    print(f"Seed(s): {unique_join(run_rows, 'seed')}")
    print(f"Test DB: {unique_join(run_rows, 'test_db')}")
    print(f"Cardinality type: {unique_join(run_rows, 'cardinality_type')}")

    print()
    print("Validation metrics (mean ± sample std):")
    print("  " + "  ".join(f"{label:>20}" for label in ("loss", "mse", "mape", "q50", "q95", "q99")))
    print("  " + "  ".join(f"{fmt_pair(summary['validation'].get(label, (None, None, 0))):>20}"
                              for label in ("loss", "mse", "mape", "q50", "q95", "q99")))

    if summary["workloads"]:
        print()
        print("Final workload metrics (mean ± sample std):")
        print("  workload                  loss                 mse                  mape                 q50                  q95                  q99                  worst_q")
        for workload, metrics in summary["workloads"].items():
            print(
                f"  {workload[:24]:24} "
                + " ".join(f"{fmt_pair(metrics[label]):>20}"
                           for label in ("loss", "mse", "mape", "q50", "q95", "q99", "worst_q"))
            )

    if summary["failed_runs"]:
        print(f"Warning: {summary['failed_runs']} requested run(s) failed or had no successful summary.")
    print(SUMMARY_END)


def format_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for column_cells in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 50)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def xlsx_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def write_xlsx(summary, path: str) -> None:
    metric_labels = [label for label, _, _ in METRICS]
    fieldnames = [
        "scope", "workload", "requested_runs", "successful_runs", "cuda_devices", "seeds",
        "test_db", "cardinality_type",
    ]
    for label in metric_labels:
        fieldnames.extend([f"{label}_mean", f"{label}_std", f"{label}_count"])

    run_rows = summary["run_rows"]
    common = {
        "requested_runs": summary["requested_runs"],
        "successful_runs": summary["successful_runs"],
        "cuda_devices": unique_join(run_rows, "cuda_device"),
        "seeds": unique_join(run_rows, "seed"),
        "test_db": unique_join(run_rows, "test_db"),
        "cardinality_type": unique_join(run_rows, "cardinality_type"),
    }

    output_rows = []
    validation_row = {**common, "scope": "validation", "workload": ""}
    for label in metric_labels:
        mean, std, count = summary["validation"].get(label, (None, None, 0))
        validation_row.update({f"{label}_mean": mean, f"{label}_std": std, f"{label}_count": count})
    output_rows.append(validation_row)

    for workload, metrics in summary["workloads"].items():
        row = {**common, "scope": "workload", "workload": workload}
        for label in metric_labels:
            mean, std, count = metrics[label]
            row.update({f"{label}_mean": mean, f"{label}_std": std, f"{label}_count": count})
        output_rows.append(row)

    output_dir = os.path.dirname(path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    workbook = Workbook()
    aggregate_sheet = workbook.active
    aggregate_sheet.title = "Aggregate"
    aggregate_sheet.append(fieldnames)
    for row in output_rows:
        aggregate_sheet.append([xlsx_value(row.get(key)) for key in fieldnames])
    format_worksheet(aggregate_sheet)

    run_details = summary["run_details"]
    if run_details:
        detail_fields = []
        seen_fields = set()
        for detail in run_details:
            for field in detail:
                if field not in seen_fields:
                    seen_fields.add(field)
                    detail_fields.append(field)
        detail_sheet = workbook.create_sheet("Run Details")
        detail_sheet.append(detail_fields)
        for detail in run_details:
            detail_sheet.append([xlsx_value(detail.get(field)) for field in detail_fields])
        format_worksheet(detail_sheet)

    workbook.save(path)
    print(f"Aggregate XLSX: {path}")


def main() -> int:
    args = parse_args()
    rows, run_details = read_workbooks(args.summary_xlsx)
    summary = aggregate(rows, args.requested_runs, run_details)
    print_summary(summary)
    write_xlsx(summary, args.output_xlsx)
    return 0 if summary["successful_runs"] > 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
