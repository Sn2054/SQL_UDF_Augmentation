#!/usr/bin/env python3
"""Plot baseline versus augmented validation q-error means for one test database."""

from __future__ import annotations

import argparse
import math
import os
import re
from pathlib import Path
from typing import Dict, Optional

os.environ.setdefault("MPLCONFIGDIR", "/tmp/polaris_matplotlib")

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

from summarize_repeated_runs import aggregate, read_workbooks


METRICS = ("q50", "q95", "q99")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("summary_xlsx", nargs="+", help="Per-run augmented summary XLSX files")
    parser.add_argument("--requested-runs", type=int, required=True)
    parser.add_argument("--test-db", required=True)
    parser.add_argument("--cardinality", required=True)
    parser.add_argument("--time-stamp", required=True)
    parser.add_argument("--baseline-xlsx", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--seed", required=True)
    parser.add_argument("--augment", required=True)
    parser.add_argument("--augment-pooling", required=True)
    parser.add_argument("--augment-refinement", required=True)
    parser.add_argument("--augment-coarse-layers", required=True)
    parser.add_argument("--augment-include-inv", required=True)
    parser.add_argument("--augment-refine-ret", required=True)
    parser.add_argument("--lambda-struct", required=True)
    return parser.parse_args()


def as_float(value: object) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def read_baseline(path: Path, test_db: str, cardinality: str) -> Optional[Dict[str, Optional[float]]]:
    if not path.exists():
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Baseline"] if "Baseline" in workbook.sheetnames else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        workbook.close()
        return None
    header_index = {str(header): index for index, header in enumerate(headers) if header is not None}
    required = {"test_db", "cardinality_type", "q50_mean", "q95_mean", "q99_mean"}
    if not required.issubset(header_index):
        workbook.close()
        return None

    result = None
    for row in rows:
        database = str(row[header_index["test_db"]] or "")
        row_cardinality = str(row[header_index["cardinality_type"]] or "")
        if database.casefold() == test_db.casefold() and row_cardinality.casefold() == cardinality.casefold():
            result = {
                metric: as_float(row[header_index[f"{metric}_mean"]])
                for metric in METRICS
            }
            break
    workbook.close()
    return result


def aggregate_augmented(args: argparse.Namespace) -> Optional[Dict[str, Optional[float]]]:
    rows, run_details = read_workbooks(args.summary_xlsx)
    summary = aggregate(rows, args.requested_runs, run_details)
    if summary["successful_runs"] == 0:
        return None
    return {
        metric: summary["validation"].get(metric, (None, None, 0))[0]
        for metric in METRICS
    }


def bar_values(values: Dict[str, Optional[float]]):
    return [np.nan if values[metric] is None else values[metric] for metric in METRICS]


def label_bars(axis, bars, values) -> None:
    for bar, value in zip(bars, values):
        if np.isnan(value):
            axis.text(
                bar.get_x() + bar.get_width() / 2,
                0,
                "N/A",
                ha="center",
                va="bottom",
                fontsize=10,
            )
        else:
            axis.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height(),
                f"{value:.4f}",
                ha="center",
                va="bottom",
                fontsize=10,
            )


def safe_filename_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "-", value).strip("-_") or "unknown"


def create_plot(
        baseline: Dict[str, Optional[float]],
        augmented: Dict[str, Optional[float]],
        args: argparse.Namespace) -> Path:
    metric_labels = [metric.upper() for metric in METRICS]
    x = np.arange(len(metric_labels))
    bar_width = 0.32
    baseline_values = bar_values(baseline)
    augmented_values = bar_values(augmented)

    figure, axis = plt.subplots(figsize=(11, 7))
    baseline_bars = axis.bar(
        x - bar_width / 2,
        baseline_values,
        width=bar_width,
        label="Baseline",
        color="#4C78A8",
        edgecolor="black",
        linewidth=0.5,
    )
    augmented_bars = axis.bar(
        x + bar_width / 2,
        augmented_values,
        width=bar_width,
        label="Augmentation",
        color="#F58518",
        edgecolor="black",
        linewidth=0.5,
    )
    label_bars(axis, baseline_bars, baseline_values)
    label_bars(axis, augmented_bars, augmented_values)

    finite_values = [
        value
        for value in baseline_values + augmented_values
        if not np.isnan(value)
    ]
    maximum = max(finite_values, default=1.0)
    axis.set_ylim(0, maximum * 1.18 if maximum > 0 else 1.0)
    axis.set_xticks(x)
    axis.set_xticklabels(metric_labels, fontsize=11)
    axis.set_xlim(-0.5, len(metric_labels) - 0.5)
    axis.set_ylabel("Q-error", fontsize=11)
    axis.grid(axis="y", linestyle="--", alpha=0.3)
    axis.set_axisbelow(True)
    axis.legend(loc="upper left", frameon=False, ncol=2)
    figure.suptitle(
        "Baseline vs Augmentation Q-error\n"
        f"Test DB: {args.test_db} | Cardinality: {args.cardinality}",
        fontsize=17,
        fontweight="bold",
        y=0.98,
    )

    settings = (
        f"SEED={args.seed} | AUGMENT={args.augment} | "
        f"AUGMENT_POOLING={args.augment_pooling} | "
        f"AUGMENT_REFINEMENT={args.augment_refinement}\n"
        f"AUGMENT_COARSE_LAYERS={args.augment_coarse_layers} | "
        f"AUGMENT_INCLUDE_INV={args.augment_include_inv} | "
        f"AUGMENT_REFINE_RET={args.augment_refine_ret} | "
        f"LAMBDA_STRUCT={args.lambda_struct}"
    )
    figure.text(
        0.5,
        0.80,
        settings,
        ha="center",
        va="center",
        fontsize=12,
        linespacing=1.5,
    )
    figure.tight_layout(rect=[0, 0, 1, 0.76])

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / (
        f"{safe_filename_part(args.test_db)}_{safe_filename_part(args.time_stamp)}.png"
    )
    figure.savefig(output_path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(figure)
    return output_path


def main() -> int:
    args = parse_args()
    baseline = read_baseline(Path(args.baseline_xlsx), args.test_db, args.cardinality)
    if baseline is None:
        print(
            "Augmented plot skipped: no baseline row for "
            f"({args.test_db}, {args.cardinality}) in {args.baseline_xlsx}"
        )
        return 0
    augmented = aggregate_augmented(args)
    if augmented is None:
        print("Augmented plot skipped: no successful augmented runs were found.")
        return 0
    output_path = create_plot(baseline, augmented, args)
    print(f"Augmented plot: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
